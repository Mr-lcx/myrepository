import openpyxl


class ReadExcle:
    def __init__(self):
        self.mydata=[] # 成绩列表

    def read(self,FileName):
        wb=openpyxl.load_workbook(FileName)
        sheet = wb.worksheets[0]  # 获取第一张表格
        maxrow=sheet.max_row
        print("行数：",maxrow)
        for row in range(2,maxrow+1):
            data=[]
            data.append(str(sheet.cell(row=row,column=1).value)) # 学号
            data.append(sheet.cell(row=row,column=2).value)# 课程号
            data.append(sheet.cell(row=row,column=3).value)# 平时成绩
            data.append(sheet.cell(row=row,column=4).value)# 期末成绩
            data.append(sheet.cell(row=row, column=5).value)# 总评
            self.mydata.append(data)
        print(self.mydata)
        return self.mydata #成绩列表

    def readstudent(self,FileName):
        wb=openpyxl.load_workbook(FileName)
        sheet = wb.worksheets[0]  # 获取第一张表格
        maxrow=sheet.max_row
        print("行数：",maxrow)
        for row in range(2,maxrow+1):
            data=[]
            data.append(str(sheet.cell(row=row,column=1).value)) # 学号
            data.append(sheet.cell(row=row,column=2).value)# 姓名
            data.append(sheet.cell(row=row,column=3).value)# 性别
            data.append(sheet.cell(row=row,column=4).value)# 密码
            data.append(sheet.cell(row=row, column=5).value)# 找回密码
            data.append(sheet.cell(row=row, column=6).value)  # 学院
            data.append(sheet.cell(row=row, column=7).value)  # 班级
            self.mydata.append(data)
        print(self.mydata)
        return self.mydata #成绩列表

    def write(self,data,fileName):
        wb = openpyxl.Workbook()  # 创建一个空的Excel 文档
        sheet = wb.create_sheet(index=0, title='成绩')  # 创建一个表格sheet
        title = ['课程号', '课程名', '学号', '姓名', '学院', '班级', '平时成绩', '期末成绩', '总评成绩']
        for i in range(1,len(title)+1):
            sheet.cell(row=1,column=i).value=title[i-1]
        for i in range(2,len(data)+2):
            sheet.cell(row=i, column=1).value = data[i-2][0]
            sheet.cell(row=i, column=2).value = data[i - 2][1]
            sheet.cell(row=i, column=3).value = data[i - 2][2]
            sheet.cell(row=i, column=4).value = data[i - 2][3]
            sheet.cell(row=i, column=5).value = data[i - 2][4]
            sheet.cell(row=i, column=6).value = data[i - 2][5]
            sheet.cell(row=i, column=7).value = data[i - 2][6]
            sheet.cell(row=i, column=8).value = data[i - 2][7]
            sheet.cell(row=i, column=9).value = data[i - 2][8]
        wb.save(fileName+'.xlsx')  # 打开或者创建Excel后都要保存